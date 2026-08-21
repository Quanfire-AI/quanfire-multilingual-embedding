"""
Reading MILPaC into a held-out evaluation set.

MILPaC — the Multilingual Indian Legal Parallel Corpus — is a small,
expert-ratified set of English legal text aligned to nine Indian
languages. It is the **evaluation front door**, the counterpart to
:mod:`.wikipedia` at the training end, and the two are deliberately kept
apart.

Two properties make it the right held-out set, and both are structural
rather than a matter of discipline:

* **Different origin from the training corpus.** The training data for
  this domain is public court judgments; MILPaC is translated legislation,
  FAQs and IP-office text. A model scored on text drawn from the same
  source it trained on is scored on its own memorisation. Holding out *by
  origin* is stronger than holding out by a random split of one corpus,
  because a random split still shares vocabulary, register and formatting
  across the boundary.

* **The alignment is the supervision.** Each row is one English unit and
  its human translation, so an anchor and its positive already exist —
  there is nothing to mine. That is why this module does **not** route
  through :mod:`.pairs`: running an aligned pair through a pair *miner*
  would be inventing structure that is already present, and worse, it
  would put the evaluation set through the same manufacturing step as the
  training set and quietly erase the origin wall.

The licence is the second reason the wall exists. MILPaC is
**CC BY-NC-SA 4.0** — non-commercial. Scoring a model against it is fair
use of an evaluation benchmark; **training a shipped adapter on it is
not**, because the adapter is a commercial artefact and NC travels into
it. So this module produces an evaluation file and nothing else, every
record carries ``license`` so the restriction is visible at the point of
use, and the training path has no way to reach it.

The corpus ships as ``.xlsx`` — three workbooks (Acts, IP, CCI-FAQ),
each with the columns ``dataset, id, src_lang, src, tgt_lang, tgt``. One
row is one parallel unit. ``openpyxl`` lives behind the optional
``milpac`` extra, imported inside the functions here, so a core install
still imports :mod:`multilingual_embedding.corpus`.
"""

from __future__ import annotations

import gzip
import json
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from multilingual_embedding.core.exceptions import MultilingualEmbeddingError, ValidationError
from multilingual_embedding.core.logging import get_logger

from .language import normalize_language_code

__all__ = [
    "MILPAC_LICENSE",
    "MILPAC_SOURCE",
    "MilpacExtractionError",
    "MilpacUnit",
    "extract_milpac",
    "iter_units",
]

_logger = get_logger(__name__)

# Recorded on every evaluation record. Kept as constants rather than
# string literals because both are load-bearing: the source is what an
# audit groups by, and the licence is the reason this text may be scored
# against but not trained on.
MILPAC_SOURCE = "milpac"

MILPAC_LICENSE = "CC BY-NC-SA 4.0"

# The two Indic targets locked for this project. MILPaC carries nine; the
# rest are skipped rather than silently pulled in, because an evaluation
# set that includes languages the model was never adapted for reports a
# recall that is really a measure of the base checkpoint's coverage.
_DEFAULT_TARGET_LANGUAGES: tuple[str, ...] = ("hi", "ta")

# The English side. MILPaC is English-centric, so this is what every
# anchor is, and a row whose source is not English is a schema surprise
# worth skipping and counting rather than trusting.
_DEFAULT_SOURCE_LANGUAGE = "en"

# The columns the workbooks are documented to carry. Matched by name off
# the header row rather than by position, so a reordered export still
# reads correctly and a renamed one fails loudly instead of reading the
# wrong column.
_REQUIRED_COLUMNS: tuple[str, ...] = ("dataset", "id", "src_lang", "src", "tgt_lang", "tgt")


class MilpacExtractionError(MultilingualEmbeddingError):
    """Raised when a MILPaC workbook cannot be read or is not what it claims."""


@dataclass(slots=True)
class MilpacUnit:
    """
    One aligned unit — an English legal passage and its translation.

    Attributes
    ----------
    dataset:
        Which MILPaC dataset the row came from (``Acts``, ``IP`` or
        ``CCI-FAQ``). Carried so an evaluation can be broken down by
        register: statute language and FAQ language score differently.

    identifier:
        The row's own id within its dataset.

    source_language, target_language:
        Normalised ISO codes. The source is English; the target is the
        Indic language whose retrieval is being measured.

    source_text:
        The English side — the anchor, the query.

    target_text:
        The translation — the positive, the passage that should be
        retrieved for it.
    """

    dataset: str

    identifier: str

    source_language: str

    target_language: str

    source_text: str

    target_text: str

    def to_pair_record(self) -> dict[str, Any]:
        """
        Reduce to a pair record the adapter's evaluation loader accepts.

        The shape is :meth:`MinedPair.to_record`'s — ``anchor`` and
        ``positive`` with provenance facets — so the file drops straight
        into ``qfme adapt --eval-pairs-file`` with no conversion. Two
        extra keys ride along that a mined pair does not carry:

        * ``kind`` is ``"parallel"`` rather than one of the mined kinds,
          because these were aligned by a translator, not read off
          document structure. It never collides with a mined kind, so a
          ``--eval-kinds`` filter can select the parallel set exactly.
        * ``source`` and ``license`` are stamped on every record.
          :meth:`MinedPair.from_record` ignores unknown keys, so this
          costs the loader nothing — but it means the non-commercial
          restriction is legible in the raw file, at the point someone
          might be tempted to feed it to ``--pairs``.

        ``language`` is the *target* language, because that is the side
        retrieval is scored on. An ``--eval-languages hi,ta`` filter
        therefore selects by the Indic language, which is what a
        per-language recall table is broken down by.
        """

        return {
            "anchor": self.source_text,
            "positive": self.target_text,
            "kind": "parallel",
            "document": (
                f"milpac-{self.dataset}-{self.source_language}-{self.target_language}"
                f"-{self.identifier}"
            ),
            "language": self.target_language,
            "source": MILPAC_SOURCE,
            "license": MILPAC_LICENSE,
        }


def _workbooks(path: Path) -> list[Path]:
    """Resolve a path to the workbooks under it, sorted for reproducibility."""

    if path.is_dir():
        found = sorted(path.glob("*.xlsx"))

        if not found:
            raise MilpacExtractionError("No .xlsx workbooks found in directory", path=str(path))

        return found

    if not path.exists():
        raise MilpacExtractionError("MILPaC workbook not found", path=str(path))

    return [path]


def _header_index(row: tuple[Any, ...]) -> dict[str, int]:
    """
    Map each required column name to its position in the header row.

    Raises loudly if a documented column is missing, because reading the
    wrong column produces a plausible evaluation file rather than an
    error — an anchor that is really an id, scored and believed.
    """

    header = {
        str(value).strip().lower(): index for index, value in enumerate(row) if value is not None
    }

    missing = [column for column in _REQUIRED_COLUMNS if column not in header]

    if missing:
        raise MilpacExtractionError(
            "Workbook is missing expected MILPaC columns",
            missing=missing,
            found=sorted(header),
        )

    return {column: header[column] for column in _REQUIRED_COLUMNS}


def _normalise(code: Any) -> str | None:
    """Normalise a language cell, returning ``None`` if it is not a code."""

    if code is None or not str(code).strip():
        return None

    try:
        return normalize_language_code(str(code))
    except ValidationError:
        return None


def iter_units(
    source: str | Path,
    *,
    languages: tuple[str, ...] = _DEFAULT_TARGET_LANGUAGES,
    datasets: tuple[str, ...] | None = None,
    source_language: str = _DEFAULT_SOURCE_LANGUAGE,
) -> Iterator[MilpacUnit]:
    """
    Stream aligned units from one workbook or a directory of them.

    Rows are filtered to the requested Indic targets and, optionally, to
    named datasets. A row whose source or target text is blank, or whose
    language cell is not a recognisable code, is skipped and counted
    rather than emitted — an evaluation set is small and every bad row in
    it distorts a recall figure more than one bad row in a training set.

    Parameters
    ----------
    source:
        A ``.xlsx`` workbook, or a directory holding several. A directory
        reads every ``*.xlsx`` under it, sorted.

    languages:
        Target Indic languages to keep. Defaults to the two this project
        adapted for; widening it measures languages the model may not
        have been trained on.

    datasets:
        If given, keep only rows whose ``dataset`` column matches one of
        these (case-insensitively). ``None`` keeps all three.

    source_language:
        The language every anchor must be in. Defaults to English, and a
        row that is not English on the source side is skipped, because
        MILPaC is English-centric and a non-English source is a schema
        surprise rather than a usable cross-lingual pair.

    Raises
    ------
    MilpacExtractionError
        If ``openpyxl`` is not installed, a workbook is unreadable, or a
        workbook does not carry the documented columns.
    """

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise MilpacExtractionError(
            "Reading MILPaC needs the 'milpac' extra (uv sync --extra milpac)",
            missing="openpyxl",
        ) from error

    wanted_languages = {normalize_language_code(code) for code in languages}

    wanted_datasets = {name.strip().lower() for name in datasets} if datasets else None

    source_code = normalize_language_code(source_language)

    produced = 0

    skipped_language = 0

    skipped_dataset = 0

    skipped_blank = 0

    skipped_source = 0

    for workbook_path in _workbooks(Path(source).expanduser()):
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as error:  # openpyxl raises a zoo of types on a bad file
            raise MilpacExtractionError(
                "Could not open MILPaC workbook",
                path=str(workbook_path),
                detail=str(error),
            ) from error

        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)

                try:
                    header_row = next(rows)
                except StopIteration:
                    continue

                columns = _header_index(header_row)

                for row in rows:
                    dataset = str(row[columns["dataset"]] or "").strip()

                    if wanted_datasets is not None and dataset.lower() not in wanted_datasets:
                        skipped_dataset += 1

                        continue

                    target_language = _normalise(row[columns["tgt_lang"]])

                    if target_language is None or target_language not in wanted_languages:
                        skipped_language += 1

                        continue

                    if _normalise(row[columns["src_lang"]]) != source_code:
                        skipped_source += 1

                        continue

                    src_text = str(row[columns["src"]] or "").strip()

                    tgt_text = str(row[columns["tgt"]] or "").strip()

                    if not src_text or not tgt_text:
                        skipped_blank += 1

                        continue

                    produced += 1

                    yield MilpacUnit(
                        dataset=dataset,
                        identifier=str(row[columns["id"]] or "").strip(),
                        source_language=source_code,
                        target_language=target_language,
                        source_text=src_text,
                        target_text=tgt_text,
                    )
        finally:
            workbook.close()

    _logger.info(
        "Read MILPaC evaluation units",
        extra={
            "units": produced,
            "skipped_other_language": skipped_language,
            "skipped_other_dataset": skipped_dataset,
            "skipped_non_english_source": skipped_source,
            "skipped_blank": skipped_blank,
        },
    )


def extract_milpac(
    source: str | Path,
    output: str | Path,
    *,
    languages: tuple[str, ...] = _DEFAULT_TARGET_LANGUAGES,
    datasets: tuple[str, ...] | None = None,
    source_language: str = _DEFAULT_SOURCE_LANGUAGE,
) -> int:
    """
    Write a MILPaC evaluation pair file and return the unit count.

    The output is JSON Lines of pair records, gzipped when the path ends
    ``.gz``. It is meant for ``qfme adapt --eval-pairs-file``, and it must
    never be passed to ``--pairs``: the licence is non-commercial and the
    whole point of a separate front door is that the evaluation set and
    the training set do not share an origin.

    Example
    -------
    ::

        extract_milpac("data/corpora/milpac/", "data/eval/milpac-hi-ta.jsonl.gz")
    """

    destination = Path(output).expanduser()

    destination.parent.mkdir(parents=True, exist_ok=True)

    opener = gzip.open if destination.suffix == ".gz" else open

    count = 0

    with opener(destination, "wt", encoding="utf-8") as handle:
        for unit in iter_units(
            source,
            languages=languages,
            datasets=datasets,
            source_language=source_language,
        ):
            handle.write(json.dumps(unit.to_pair_record(), ensure_ascii=False) + "\n")

            count += 1

    _logger.info(
        "Wrote MILPaC evaluation set",
        extra={"path": str(destination), "units": count},
    )

    return count
